import re
import os
import datetime
import requests
import openpyxl
from bs4 import BeautifulSoup

class WebScraperFileDownloader:

    def get_links(link):

        link = 'https://spo-13.mskobr.ru/uchashimsya/raspisanie-kanikuly'
        response = requests.get(link)
        soup = BeautifulSoup(response.text, 'html.parser')

        links = []
        for a in soup.find_all('a', href=True):
            url = a['href']
            parts = re.split(r'/', url)
            if 'files' in parts and 'attach_files' in parts:
                links.append(url)

        i = len(links) - 1
        url = f"https://spo-13.mskobr.ru/{links[i]}"
        return url


    def download_file(self, url, name):

        folder_path = f"src\\dataname"
        os.makedirs(folder_path, exist_ok=True)

        try:

            response = requests.get(url)

            if response.status_code == 200:
                file_name = f"{name}.xlsx"
                file_path = os.path.join(folder_path, file_name)
                with open(file_path, "wb") as file:
                    file.write(response.content)
                print(f"File downloaded successfully. Time: {datetime.datetime.now().time()} Status Code: {response.status_code}")
            else:
                print(f"Error downloading file: {url}\nTime: {datetime.datetime.now().time()} Status Code: {response.status_code}")

        except PermissionError:
            print("Закройте файл.")

class ExcelDataParser():

    def readDataShankursky(self, path, sheet):
        
        wb = openpyxl.load_workbook(path)
        sheet = wb[sheet]
        
        ranges = [
            (4, 12),
            (13, 20),
            (21, 28),
            (21, 28),
            (29, 36),
            (37, 44)
        ]    

        # Общее расписание из всех чанков
        timetable = []
                                                                                # работа с конкретной группой. range(a, b), где: a - порядковое число столбца, b - порядковое число строки 
        for column in range(3, 21, 2):
                                                                                # проход по всем диапазонам ячеек
            for rangeStart, rangeEnd in ranges:
                                                                                # отдельный список (чанк) для каждого дня
                chunk = []
                                                                                # работа с конкретным диапазоном
                for index in range(rangeStart, rangeEnd + 1):
                                                                                # Получаем значения из ячеек 
                    cellValue = sheet.cell(row=index, column=column).value
                    chunk.append(cellValue)
                                                                                # Добавляем каждый список (чанк) в общий список    
                timetable.append(chunk)

        return timetable


#downloader = WebScraperFileDownloader()
#url = downloader.get_links()
#name = "test-Shankursky"
#downloader.download_file(url, name)

#parser = ExcelDataParser()
#wb = openpyxl.load_workbook(f"src\\dataname\\{name}\\{name}.xlsx")
#sheet = wb["Лист1"]

#reader = ExcelDataParser()
# [-index], где index - обратный отсчет каждого дня (чанк) расписания начиная с правого нижнего угла, снизу вверх, переход на новый столбец - снова снизу вверх
#lessons = reader.readDataShankursky(f"src\\dataname\\{name}\\{name}.xlsx", 'Лист1')[-10]
#lessons = [lesson if lesson is not None else "Нет предметов" for lesson in lessons]

#mondayLessons = '\n'.join([f"▫️ {lesson}" for i, lesson in enumerate(lessons)])

#print(mondayLessons)